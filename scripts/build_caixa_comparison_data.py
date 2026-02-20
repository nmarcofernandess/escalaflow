#!/usr/bin/env python3
"""Gera arquivo unico de referencia da Rita (escala + pessoas por horario)."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def build_rita_reference(escala_xlsx: Path, dom_folgas_xlsx: Path, out_path: Path) -> None:
    wb = load_workbook(str(escala_xlsx), data_only=True)

    name_map = {
        "CLEONICE": "Mão de Obra - Cleonice",
        "GABRIEL": "Mão de Obra - Gabriel",
        "ANA JULIA": "Mão de Obra - Ana Julia",
        "YASMIN": "Mão de Obra - Yasmin",
        "MAYUMI": "Mão de Obra - Mayumi",
        "HELOISA": "Mão de Obra - Heloisa",
    }
    day_map = {
        "SEGUNDA": "2026-02-09",
        "TERÇA": "2026-02-10",
        "TERCA": "2026-02-10",
        "QUARTA": "2026-02-11",
        "QUINTA": "2026-02-12",
        "SEXTA": "2026-02-13",
        "SÁBADO": "2026-02-14",
        "SABADO": "2026-02-14",
    }

    def to_min(hhmm: str) -> int:
        h, m = map(int, hhmm.split(":"))
        return h * 60 + m

    def parse_scale8(ws) -> dict:
        start = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().upper() == "ESCALA 8":
                start = r
                break
        if not start:
            return {}

        out = {}
        for r in range(start + 2, start + 9):
            d = ws.cell(r, 1).value
            if not isinstance(d, str):
                continue
            du = d.strip().upper()
            if du not in day_map:
                continue

            day = day_map[du]
            ent = ws.cell(r, 2).value
            i1 = ws.cell(r, 3).value
            i2 = ws.cell(r, 4).value
            sai = ws.cell(r, 5).value

            def norm(v):
                if v is None:
                    return None
                s = str(v).strip()
                return None if s in ("", "0:00") else s

            ent = norm(ent)
            i1 = norm(i1)
            i2 = norm(i2)
            sai = norm(sai)

            if ent is None or sai is None:
                out[day] = {"inicio": None, "fim": None, "almoco": None, "minutos": 0}
            else:
                lunch = 0
                alm = None
                if i1 and i2:
                    alm = f"{i1}-{i2}"
                    lunch = to_min(i2) - to_min(i1)
                mins = (to_min(sai) - to_min(ent)) - lunch
                out[day] = {"inicio": ent, "fim": sai, "almoco": alm, "minutos": mins}

        return out

    days = ["2026-02-09", "2026-02-10", "2026-02-11", "2026-02-12", "2026-02-13", "2026-02-14"]
    aloc = {d: {} for d in days}
    horas_dia = {}
    horas_sem = {}

    for nome, sheet in name_map.items():
        esc = parse_scale8(wb[sheet])
        horas_dia[nome] = {
            "SEG": esc["2026-02-09"]["minutos"],
            "TER": esc["2026-02-10"]["minutos"],
            "QUA": esc["2026-02-11"]["minutos"],
            "QUI": esc["2026-02-12"]["minutos"],
            "SEX": esc["2026-02-13"]["minutos"],
            "SAB": esc["2026-02-14"]["minutos"],
        }
        horas_sem[nome] = sum(horas_dia[nome].values())
        for d in days:
            aloc[d][nome] = esc[d]

    # Deriva curva de pessoas por slot (30min) a partir da escala real da Rita.
    def day_label(iso_day: str) -> str:
        return {
            "2026-02-09": "SEG",
            "2026-02-10": "TER",
            "2026-02-11": "QUA",
            "2026-02-12": "QUI",
            "2026-02-13": "SEX",
            "2026-02-14": "SAB",
        }[iso_day]

    demanda_por_horario = []
    for d in days:
        by_name = aloc[d]
        for t in range(8 * 60, 20 * 60, 30):
            cov = 0
            for alloc in by_name.values():
                ini = alloc.get("inicio")
                fim = alloc.get("fim")
                if not ini or not fim:
                    continue
                s = to_min(ini)
                e = to_min(fim)
                if not (s <= t and e >= t + 30):
                    continue
                alm = alloc.get("almoco")
                if alm:
                    a, b = [to_min(x.strip()) for x in alm.split("-")]
                    if a <= t and b >= t + 30:
                        continue
                cov += 1
            if cov > 0:
                demanda_por_horario.append(
                    {
                        "dia_semana": day_label(d),
                        "hora_inicio": f"{t // 60:02d}:{t % 60:02d}",
                        "hora_fim": f"{(t + 30) // 60:02d}:{(t + 30) % 60:02d}",
                        "min_pessoas": cov,
                    }
                )

    ref = {
        "metadata": {"setor": "CAIXA", "periodo": {"inicio": "2026-02-09", "fim": "2026-02-14"}},
        "source": f"{escala_xlsx.name} (ESCALA 8) + {dom_folgas_xlsx.name}",
        "colaboradores": list(name_map.keys()),
        "referencia_demanda_por_horario": demanda_por_horario,
        "ground_truth": {
            "horas_por_dia": horas_dia,
            "horas_semanais_verificacao": horas_sem,
            "alocacoes_por_dia": aloc,
        },
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(ref, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--escala-xlsx", default="/Users/marcofernandes/Downloads/escalas_trabalho-GPT.xlsx")
    parser.add_argument("--dom-folgas-xlsx", default="/Users/marcofernandes/Downloads/DOM E FOLGAS - CAIXA.xlsx")
    parser.add_argument("--out-rita", default="data/comparacao/caixa_rita_referencia.json")
    args = parser.parse_args()

    build_rita_reference(Path(args.escala_xlsx), Path(args.dom_folgas_xlsx), Path(args.out_rita))
    print(f"OK rita: {args.out_rita}")


if __name__ == "__main__":
    main()
