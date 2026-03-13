#!/usr/bin/env python3
"""
extract-xlsx-seed2.py
Reads the master Excel schedule and produces structured JSON for EscalaFlow seed2.

Output: /Users/marcofernandes/escalaflow/data/seed2/
  - colaboradores.json        All employees from PARÂMETROS
  - setores.json              Sector metadata
  - demanda_{key}.json        Demand per slot per day (per sector)
  - horarios_{key}.json       Individual weekly schedules (per sector)
  - diagnostico.json          Cross-references and summary
"""

import json
import os
import datetime
from collections import defaultdict

import openpyxl

# ─── Config ──────────────────────────────────────────────────────────
XLSX = "/Users/marcofernandes/Documents/Horario de trabalho padrao NOVO -2026 - REVISAO DE ESCALA.xlsx"
OUT  = "/Users/marcofernandes/escalaflow/data/seed2"

DAYS = {"SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"}
DAY_ALIAS = {"segunda": "SEG", "terça": "TER", "quarta": "QUA",
             "quinta": "QUI", "sexta": "SEX", "sábado": "SAB",
             "sabado": "SAB", "domingo": "DOM"}
DAY_ORDER = ["SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"]

# Noise: values in name column that are NOT real employee names
NOISE = {
    "ESC A", "ESC B", "ESCALA 1", "ESCALA 2", "ESCALA A", "ESCALA B",
    "CAIXAS", "LIMPEZA", "ENTREGA", "ROTISSERIA", "REPOSIÇÃO", "ESCRITORIO",
    "EMBALAGEM", "PADARIA", "CONFEITARIA", "SALGADERIA", "PANIFICAÇÃO",
    "DEPÓSITO", "AÇOUGUE", "PRODUÇÃO", "SAC", "PAD ATENDIMENTO",
    "NATAL", "SUPORTE",
    "HORARIO ESPECIAL", "ESCALA",
}

# Known section names that can appear as labels in multi-section sheets
SECTION_NAMES = {
    "LIMPEZA", "REPOSIÇÃO", "ENTREGA", "EMBALAGEM", "ROTISSERIA",
    "PANIFICAÇÃO", "CONFEITARIA", "SALGADERIA", "PRODUÇÃO",
    "CAIXAS", "PADARIA", "DEPÓSITO", "ESCRITORIO", "AÇOUGUE",
}

# Standard schedule sheets (one first-week block each)
SCHED = [
    ("caixas",          "CAIXAS -2026- ESC 5x2",       "CAIXAS"),
    ("pad_atendimento", "PAD ATENDIMENTO -2026- ESC",   "PAD ATENDIMENTO"),
    ("reposicao",       "REPOSIÇÃO-2025",               "REPOSIÇÃO"),
    ("acougue",         "AÇOUGUE-2025",                 "AÇOUGUE"),
    ("entrega",         "ENTREGA",                      "ENTREGA"),
    ("limpeza",         "LIMPEZA",                      "LIMPEZA"),
    ("rotisseria",      "ROTISSERIA ",                  "ROTISSERIA"),
    ("deposito",        "DEPÓSITO",                     "DEPÓSITO"),
    ("escritorio",      "ESCRITÓRIO",                   "ESCRITÓRIO"),
]

# Multi-section sheets
MULTI = [
    ("loja",    "LOJA",     "LOJA"),
    ("padaria", "PADARIA",  "PADARIA"),
]


# ─── Helpers ──────────────────────────────────────────────────────────
def ts(t):
    return t.strftime("%H:%M") if isinstance(t, datetime.time) else str(t)

def add30(s):
    h, m = map(int, s.split(":"))
    m += 30
    if m >= 60: h += 1; m -= 60
    return f"{h:02d}:{m:02d}"

def clean(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def norm_day(v):
    if v is None: return None
    s = str(v).strip()
    if s.upper() in DAYS: return s.upper()
    low = s.lower()
    for k, d in DAY_ALIAS.items():
        if low.startswith(k[:3]): return d
    return None

def is_noise(name):
    """Check if a 'name' is actually metadata/section label, not a person."""
    if not name: return True
    up = name.strip().upper()
    if up in NOISE: return True
    if up.startswith("DE ") and "20" in up: return True  # date ranges
    if up.startswith("HORARIO"): return True
    if up.startswith("ESCALA"): return True
    if up.startswith("SEMANA"): return True
    if "FUNCIONÁRIO" in up or "NECESSÁRIO" in up: return True
    if up.startswith("SUB "): return True  # "sub priscila" = annotation
    # Single digit or 0
    try:
        float(up)
        return True
    except ValueError:
        pass
    return False


# ─── Time header detection ───────────────────────────────────────────
def find_time_header(ws, max_row=15):
    for r in range(1, max_row + 1):
        for start in [2, 3]:
            if isinstance(ws.cell(r, start).value, datetime.time):
                end = start
                for c in range(start, start + 35):
                    if isinstance(ws.cell(r, c).value, datetime.time):
                        end = c
                    else:
                        break
                if end - start + 1 >= 10:
                    return r, start, end
    return None, None, None

def build_tmap(ws, row, cs, ce):
    m = {}
    for c in range(cs, ce + 1):
        v = ws.cell(row, c).value
        if isinstance(v, datetime.time):
            m[c] = ts(v)
    return m


# ─── Extract one weekly block (SEG..SAB/DOM) ─────────────────────────
def extract_week(ws, ncol, tmap, start, limit=300):
    """
    From `start` row, find and extract one complete week (SEG through SAB or DOM).
    For each day, collects employee rows: name (may be None) + sigla slots.
    Stops when a day repeats.
    Returns (days_data, end_row).
    """
    tcols = sorted(tmap.keys())
    days = {}
    cur_day = None
    cur_emps = []
    seen = []
    r = start

    while r <= min(ws.max_row, start + limit):
        val = ws.cell(r, ncol).value
        day = norm_day(val)

        if day is not None:
            # Verify it's a header: should have numbers in time cols
            has_num = any(isinstance(ws.cell(r, c).value, (int, float)) for c in tcols[:5])
            if has_num or day not in seen:
                if day in seen:
                    # Second occurrence = new block, stop
                    if cur_day and cur_emps:
                        days[cur_day] = cur_emps
                    return days, r
                if cur_day is not None:
                    days[cur_day] = cur_emps
                cur_day = day
                cur_emps = []
                seen.append(day)
                r += 1
                continue

        if cur_day is None:
            r += 1
            continue

        # Collect sigla slots
        name = clean(val)
        slots = {}
        for c in tcols:
            v = ws.cell(r, c).value
            if v is not None and isinstance(v, str) and v.strip() and v.strip() != '0':
                slots[c] = v.strip()

        if slots:
            cur_emps.append({"nome": name, "slots": slots})
        elif name and name != '0' and not is_noise(name):
            cur_emps.append({"nome": name, "slots": {}})

        r += 1

    if cur_day and cur_emps:
        days[cur_day] = cur_emps
    return days, r


# ─── Multi-section extraction (LOJA, PADARIA) ────────────────────────
def extract_sections(ws, ncol, tmap):
    """
    Scan sheet for named sections. Section labels appear in col C (or col B)
    before the first SEG of that section. May be on the same row as 'ESC A',
    or on a separate row.

    Returns: [(section_name, days_data)]
    """
    tcols = sorted(tmap.keys())
    cs, ce = tcols[0], tcols[-1]
    sections = []
    label = None
    r = 1

    while r <= min(ws.max_row, 500):
        # Check for new time header
        if isinstance(ws.cell(r, cs).value, datetime.time):
            new = build_tmap(ws, r, cs, ce)
            if len(new) >= 10:
                tmap = new
                tcols = sorted(tmap.keys())
            r += 1
            continue

        # Look for section labels in col C and col B
        for cc in [3, 2]:
            v = clean(ws.cell(r, cc).value)
            if not v or isinstance(ws.cell(r, cc).value, (int, float, datetime.time)):
                continue
            if norm_day(v) is not None:
                continue
            up = v.strip().upper()
            if up in {"ESC A", "ESC B", "ESCALA A", "ESCALA B"}:
                continue
            # Accept known section names explicitly
            if up in SECTION_NAMES:
                label = v.strip()
                break
            # Accept other strings that look like section names
            if len(v) > 3 and len(v) < 25 and "20" not in v and "FUNCIONÁRIO" not in up:
                label = v.strip()
                break

        # Check for SEG
        if norm_day(ws.cell(r, ncol).value) == "SEG":
            # Before extracting, look back up to 5 rows for a section label we might have missed
            if label is None:
                for back in range(r - 1, max(0, r - 6), -1):
                    for cc in [3, 2]:
                        v = clean(ws.cell(back, cc).value)
                        if not v or isinstance(ws.cell(back, cc).value, (int, float, datetime.time)):
                            continue
                        up = v.strip().upper()
                        if up in SECTION_NAMES:
                            label = v.strip()
                            break
                        if norm_day(v) is None and up not in {"ESC A", "ESC B", "ESCALA A", "ESCALA B"}:
                            if len(v) > 3 and len(v) < 25 and "20" not in v and "FUNCIONÁRIO" not in up:
                                label = v.strip()
                                break
                    if label:
                        break

            data, end = extract_week(ws, ncol, tmap, r, limit=150)
            if data and len(data) >= 3:
                sections.append((label or "UNNAMED", data))
                label = None
            r = end
            continue

        r += 1

    return sections


# ─── Build outputs from extracted data ───────────────────────────────
def build_outputs(tmap, days_data, sigla_name):
    tcols = sorted(tmap.keys())

    # Demand
    demand = {}
    for day in DAY_ORDER:
        if day not in days_data:
            continue
        emps = days_data[day]
        slots = []
        for c in tcols:
            count = sum(1 for e in emps if c in e["slots"] and not is_noise(e.get("nome")))
            slots.append({
                "hora_inicio": tmap[c],
                "hora_fim": add30(tmap[c]),
                "min_pessoas": count
            })
        demand[day] = slots

    # Individual schedules
    emp_map = defaultdict(lambda: {"sigla": None, "days": defaultdict(list)})
    for day, emps in days_data.items():
        for e in emps:
            name = e["nome"]
            # Skip noise
            if is_noise(name) and not e["slots"]:
                continue
            if is_noise(name) and e["slots"]:
                # Name is noise but has slots — use sigla to find real name
                first_sigla = next(iter(e["slots"].values()))
                name = sigla_name.get(first_sigla, first_sigla)
            if not name and e["slots"]:
                first_sigla = next(iter(e["slots"].values()))
                name = sigla_name.get(first_sigla, first_sigla)
            if not name:
                continue

            for c, sigla in e["slots"].items():
                t = tmap.get(c)
                if t:
                    emp_map[name]["days"][day].append(t)
                    if emp_map[name]["sigla"] is None:
                        emp_map[name]["sigla"] = sigla

    schedules = []
    for name, data in emp_map.items():
        if is_noise(name):
            continue
        por_dia = {}
        total = 0
        for day in DAY_ORDER:
            times = data["days"].get(day, [])
            if times:
                times.sort()
                por_dia[day] = {
                    "entrada": times[0],
                    "saida": add30(times[-1]),
                    "slots_trabalho": len(times),
                    "horas_brutas": len(times) * 0.5
                }
                total += len(times)
        if total > 0:
            schedules.append({
                "nome": name,
                "sigla": data["sigla"] or "?",
                "por_dia": por_dia,
                "total_slots_semana": total,
                "horas_brutas_semana": total * 0.5
            })

    return demand, schedules


# ─── Extract colaboradores from PARÂMETROS ───────────────────────────
def extract_colabs(wb):
    ws = wb["PARÂMETROS"]
    colabs = []
    seen = set()

    # Primary: N(14), P(16), Q(17), R(18)
    for r in range(8, 70):
        nome = clean(ws.cell(r, 14).value)
        cargo = clean(ws.cell(r, 16).value)
        sigla = clean(ws.cell(r, 17).value)
        setor = clean(ws.cell(r, 18).value)
        if cargo and sigla and sigla not in seen:
            seen.add(sigla)
            colabs.append({"nome": nome, "cargo": cargo, "sigla": sigla, "setor": setor})

    # Secondary: B(2), D(4), E(5), F(6)
    for r in range(8, 70):
        raw_nome = ws.cell(r, 2).value
        # Skip numeric values in name column (e.g. row counters)
        if isinstance(raw_nome, (int, float)):
            continue
        nome = clean(raw_nome)
        cargo = clean(ws.cell(r, 4).value)
        sigla = clean(ws.cell(r, 5).value)
        setor = clean(ws.cell(r, 6).value)
        if cargo and sigla:
            if sigla not in seen:
                seen.add(sigla)
                colabs.append({"nome": nome, "cargo": cargo, "sigla": sigla, "setor": setor})
            elif nome:
                for c in colabs:
                    if c["sigla"] == sigla and not c["nome"]:
                        c["nome"] = nome

    print(f"  {len(colabs)} colaboradores from PARÂMETROS")
    return colabs


# ─── Write helper ────────────────────────────────────────────────────
def wjson(name, data):
    with open(os.path.join(OUT, name), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ─── Main ────────────────────────────────────────────────────────────
def main():
    print(f"Loading: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    os.makedirs(OUT, exist_ok=True)

    # 1. Colaboradores
    print("\n=== Colaboradores ===")
    colabs = extract_colabs(wb)
    wjson("colaboradores.json", colabs)

    sigla_name = {c["sigla"]: c["nome"] for c in colabs if c["nome"] and c["sigla"]}

    # Track all names found in schedules
    all_sched_names = set()
    all_setores = []
    diag = {}

    # 2. Standard sheets
    print("\n=== Standard Sheets ===")
    for key, sheet, display in SCHED:
        print(f"\n  [{display}] sheet='{sheet}'")
        if sheet not in wb.sheetnames:
            print(f"    SKIP: not found")
            diag[key] = "not_found"
            continue

        ws = wb[sheet]
        tr, cs, ce = find_time_header(ws)
        if tr is None:
            print(f"    SKIP: no time header")
            diag[key] = "no_time_header"
            continue

        tmap = build_tmap(ws, tr, cs, ce)
        ncol = 1 if cs == 2 else 2  # name col = A if times start at B

        data, _ = extract_week(ws, ncol, tmap, tr + 1)
        if not data:
            print(f"    SKIP: no days")
            diag[key] = "no_days"
            continue

        demand, schedules = build_outputs(tmap, data, sigla_name)
        for s in schedules:
            all_sched_names.add(s["nome"])

        times = set()
        for ds in demand.values():
            for sl in ds:
                if sl["min_pessoas"] > 0:
                    times.add(sl["hora_inicio"])
                    times.add(sl["hora_fim"])
        h_open = min(times) if times else "08:00"
        h_close = max(times) if times else "18:00"

        all_setores.append({
            "nome": display, "sheet_name": sheet,
            "colaboradores_count": len(schedules),
            "hora_abertura": h_open, "hora_fechamento": h_close,
        })

        wjson(f"demanda_{key}.json", {"setor": display, "demanda_por_dia": demand})
        wjson(f"horarios_{key}.json", {"setor": display, "horarios": schedules})

        print(f"    Days: {list(data.keys())}  Employees: {len(schedules)}  Hours: {h_open}-{h_close}")
        for s in schedules:
            ds = ",".join(d for d in DAY_ORDER if d in s["por_dia"])
            print(f"      {s['nome']:22s} {s['sigla']:8s} {ds:27s} {s['horas_brutas_semana']:.1f}h")

        diag[key] = {"ok": True, "employees": len(schedules), "days": list(data.keys())}

    # 3. Multi-section sheets
    print("\n=== Multi-Section Sheets ===")
    for key, sheet, display in MULTI:
        print(f"\n  [{display}] sheet='{sheet}'")
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        tr, cs, ce = find_time_header(ws)
        if tr is None:
            continue

        tmap = build_tmap(ws, tr, cs, ce)
        ncol = 1 if cs == 2 else 2

        sections = extract_sections(ws, ncol, tmap)
        print(f"    Found {len(sections)} raw sections")

        seen_keys = {}
        for sect_name, data in sections:
            skey = sect_name.strip().upper()
            if skey in seen_keys:
                print(f"    SKIP dup: {sect_name}")
                continue
            seen_keys[skey] = True

            demand, schedules = build_outputs(tmap, data, sigla_name)
            for s in schedules:
                all_sched_names.add(s["nome"])

            times = set()
            for ds in demand.values():
                for sl in ds:
                    if sl["min_pessoas"] > 0:
                        times.add(sl["hora_inicio"])
                        times.add(sl["hora_fim"])
            h_open = min(times) if times else "08:00"
            h_close = max(times) if times else "18:00"

            full = f"{display} - {sect_name.strip()}" if skey != display.upper() else display
            safe_key = f"{key}_{skey.lower().replace(' ','_').replace('ã','a').replace('ç','c').replace('é','e').replace('ó','o').replace('ú','u')}" if skey != display.upper() else key

            all_setores.append({
                "nome": full, "sheet_name": sheet,
                "sub_section": sect_name.strip(),
                "colaboradores_count": len(schedules),
                "hora_abertura": h_open, "hora_fechamento": h_close,
            })

            wjson(f"demanda_{safe_key}.json", {"setor": full, "demanda_por_dia": demand})
            wjson(f"horarios_{safe_key}.json", {"setor": full, "horarios": schedules})

            print(f"    {full}: {len(schedules)} emps, {len(demand)} days, {h_open}-{h_close}")
            for s in schedules:
                ds = ",".join(d for d in DAY_ORDER if d in s["por_dia"])
                print(f"      {s['nome']:22s} {s['sigla']:8s} {ds:27s} {s['horas_brutas_semana']:.1f}h")

    # 4. Setores
    wjson("setores.json", all_setores)
    print(f"\n=== Summary ===")
    print(f"  {len(all_setores)} sectors written to setores.json")

    # 5. Diagnostics
    param_names = {c["nome"].strip().upper() for c in colabs if c["nome"]}
    sched_upper = {n.strip().upper() for n in all_sched_names if n}

    only_param = sorted(param_names - sched_upper)
    only_sched = sorted(sched_upper - param_names)

    diag_out = {
        "total_colaboradores_parametros": len(colabs),
        "total_employees_in_schedules": len(all_sched_names),
        "total_setores": len(all_setores),
        "sheets": diag,
        "setores": [{"nome": s["nome"], "count": s["colaboradores_count"],
                     "horario": f"{s['hora_abertura']}-{s['hora_fechamento']}"} for s in all_setores],
        "in_parametros_not_schedules": only_param,
        "in_schedules_not_parametros": only_sched,
    }
    wjson("diagnostico.json", diag_out)

    print(f"  PARÂMETROS: {len(colabs)} | In schedules: {len(all_sched_names)}")
    if only_param:
        print(f"  Only PARÂMETROS ({len(only_param)}): {', '.join(only_param)}")
    if only_sched:
        print(f"  Only schedules ({len(only_sched)}): {', '.join(only_sched)}")
    print(f"\n  Output: {OUT}")


if __name__ == "__main__":
    main()
